"""
Excel Reader Utility
Provides a class for reading and processing Excel files (.xlsx, .xls)
"""
import logging
from typing import List, Dict, Any, Optional, Union
from pathlib import Path
import pandas as pd
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

logger = logging.getLogger(__name__)


class ExcelReader:
    """
    A class to open and read Excel files with support for multiple sheets,
    data filtering, and conversion to various formats.
    
    Example usage:
        # Basic usage
        reader = ExcelReader('data.xlsx')
        data = reader.read_sheet('Sheet1')
        
        # Read specific columns
        data = reader.read_sheet('Sheet1', columns=['Name', 'Email', 'Phone'])
        
        # Read with row limits
        data = reader.read_sheet('Sheet1', max_rows=100)
        
        # Get all sheet names
        sheets = reader.get_sheet_names()
        
        # Read all sheets
        all_data = reader.read_all_sheets()
    """
    
    def __init__(self, file_path: str, engine: str = 'openpyxl'):
        """
        Initialize the ExcelReader with a file path.
        
        Args:
            file_path: Path to the Excel file (.xlsx or .xls)
            engine: Pandas engine to use ('openpyxl' for .xlsx, 'xlrd' for .xls)
        
        Raises:
            FileNotFoundError: If the file doesn't exist
            ValueError: If the file is not a valid Excel file
        """
        self.file_path = Path(file_path)
        self.engine = engine
        
        if not self.file_path.exists():
            raise FileNotFoundError(f"Excel file not found: {file_path}")
        
        if not self.file_path.suffix.lower() in ['.xlsx', '.xls', '.xlsm']:
            raise ValueError(f"Invalid Excel file format: {self.file_path.suffix}")
        
        # Adjust engine based on file extension
        if self.file_path.suffix.lower() == '.xls':
            self.engine = 'xlrd'
        elif self.file_path.suffix.lower() in ['.xlsx', '.xlsm']:
            self.engine = 'openpyxl'
        
        self._workbook = None
        self._excel_file = None
        
        logger.info(f"Initialized ExcelReader for file: {file_path}")
    
    def __enter__(self):
        """Context manager entry"""
        return self
    
    def __exit__(self, exc_type, exc_val, exc_tb):
        """Context manager exit - cleanup resources"""
        self.close()
    
    def get_sheet_names(self) -> List[str]:
        """
        Get all sheet names in the Excel file.
        
        Returns:
            List of sheet names
        """
        try:
            excel_file = pd.ExcelFile(self.file_path, engine=self.engine)
            sheet_names = excel_file.sheet_names
            excel_file.close()
            return sheet_names
        except Exception as e:
            logger.error(f"Error getting sheet names: {e}")
            raise
    
    def read_sheet(
        self, 
        sheet_name: Union[str, int] = 0,
        columns: Optional[List[str]] = None,
        skip_rows: int = 0,
        max_rows: Optional[int] = None,
        header_row: int = 0,
        as_dict: bool = True
    ) -> Union[List[Dict[str, Any]], pd.DataFrame]:
        """
        Read data from a specific sheet.
        
        Args:
            sheet_name: Sheet name (str) or index (int, 0-based)
            columns: List of column names to read (None = all columns)
            skip_rows: Number of rows to skip from the top
            max_rows: Maximum number of data rows to read (None = all)
            header_row: Row index to use as column headers (0-based)
            as_dict: Return as list of dictionaries (True) or pandas DataFrame (False)
        
        Returns:
            List of dictionaries or pandas DataFrame with the sheet data
        """
        try:
            # Read the Excel sheet
            df = pd.read_excel(
                self.file_path,
                sheet_name=sheet_name,
                engine=self.engine,
                skiprows=skip_rows,
                nrows=max_rows,
                header=header_row,
                usecols=columns
            )
            
            logger.info(f"Read sheet '{sheet_name}': {len(df)} rows, {len(df.columns)} columns")
            
            if as_dict:
                # Convert to list of dictionaries
                return df.to_dict('records')
            else:
                return df
                
        except Exception as e:
            logger.error(f"Error reading sheet '{sheet_name}': {e}")
            raise
    
    def read_all_sheets(
        self, 
        as_dict: bool = True
    ) -> Dict[str, Union[List[Dict[str, Any]], pd.DataFrame]]:
        """
        Read all sheets in the Excel file.
        
        Args:
            as_dict: Return each sheet as list of dictionaries (True) or DataFrame (False)
        
        Returns:
            Dictionary mapping sheet names to their data
        """
        try:
            sheet_names = self.get_sheet_names()
            all_data = {}
            
            for sheet_name in sheet_names:
                all_data[sheet_name] = self.read_sheet(
                    sheet_name=sheet_name,
                    as_dict=as_dict
                )
            
            logger.info(f"Read all {len(all_data)} sheets from file")
            return all_data
            
        except Exception as e:
            logger.error(f"Error reading all sheets: {e}")
            raise
    
    def read_range(
        self,
        sheet_name: Union[str, int] = 0,
        start_row: int = 0,
        end_row: Optional[int] = None,
        start_col: int = 0,
        end_col: Optional[int] = None
    ) -> List[List[Any]]:
        """
        Read a specific range of cells from a sheet.
        
        Args:
            sheet_name: Sheet name or index
            start_row: Starting row index (0-based)
            end_row: Ending row index (None = to end)
            start_col: Starting column index (0-based)
            end_col: Ending column index (None = to end)
        
        Returns:
            List of lists containing cell values
        """
        try:
            df = pd.read_excel(
                self.file_path,
                sheet_name=sheet_name,
                engine=self.engine,
                header=None  # Don't use first row as header
            )
            
            # Apply slicing
            if end_row is None:
                end_row = len(df)
            if end_col is None:
                end_col = len(df.columns)
            
            subset = df.iloc[start_row:end_row, start_col:end_col]
            
            return subset.values.tolist()
            
        except Exception as e:
            logger.error(f"Error reading range: {e}")
            raise
    
    def get_cell_value(
        self,
        sheet_name: Union[str, int] = 0,
        row: int = 0,
        col: int = 0
    ) -> Any:
        """
        Get value from a specific cell.
        
        Args:
            sheet_name: Sheet name or index
            row: Row index (0-based)
            col: Column index (0-based)
        
        Returns:
            Cell value
        """
        try:
            df = pd.read_excel(
                self.file_path,
                sheet_name=sheet_name,
                engine=self.engine,
                header=None
            )
            
            return df.iloc[row, col]
            
        except Exception as e:
            logger.error(f"Error getting cell value at [{row}, {col}]: {e}")
            raise
    
    def get_workbook(self):
        """
        Get the openpyxl workbook object for advanced operations.
        Only works with .xlsx files.
        
        Returns:
            openpyxl Workbook object
        """
        if self.engine != 'openpyxl':
            raise ValueError("Workbook access only available for .xlsx files with openpyxl engine")
        
        if self._workbook is None:
            self._workbook = load_workbook(self.file_path, read_only=True, data_only=True)
        
        return self._workbook
    
    def get_worksheet(self, sheet_name: Union[str, int] = 0) -> Worksheet:
        """
        Get an openpyxl worksheet object for advanced operations.
        
        Args:
            sheet_name: Sheet name or index
        
        Returns:
            openpyxl Worksheet object
        """
        workbook = self.get_workbook()
        
        if isinstance(sheet_name, int):
            sheet_names = workbook.sheetnames
            if 0 <= sheet_name < len(sheet_names):
                sheet_name = sheet_names[sheet_name]
            else:
                raise IndexError(f"Sheet index {sheet_name} out of range")
        
        return workbook[sheet_name]
    
    def get_sheet_info(self, sheet_name: Union[str, int] = 0) -> Dict[str, Any]:
        """
        Get metadata information about a sheet.
        
        Args:
            sheet_name: Sheet name or index
        
        Returns:
            Dictionary with sheet information (rows, columns, dimensions)
        """
        try:
            df = pd.read_excel(
                self.file_path,
                sheet_name=sheet_name,
                engine=self.engine
            )
            
            info = {
                'sheet_name': sheet_name if isinstance(sheet_name, str) else self.get_sheet_names()[sheet_name],
                'total_rows': len(df),
                'total_columns': len(df.columns),
                'column_names': df.columns.tolist(),
                'has_data': not df.empty,
                'memory_usage': df.memory_usage(deep=True).sum()
            }
            
            return info
            
        except Exception as e:
            logger.error(f"Error getting sheet info: {e}")
            raise
    
    def search(
        self,
        search_value: Any,
        sheet_name: Union[str, int] = 0,
        column: Optional[str] = None
    ) -> List[Dict[str, Any]]:
        """
        Search for rows containing a specific value.
        
        Args:
            search_value: Value to search for
            sheet_name: Sheet name or index
            column: Specific column to search in (None = search all columns)
        
        Returns:
            List of matching rows as dictionaries
        """
        try:
            df = pd.read_excel(
                self.file_path,
                sheet_name=sheet_name,
                engine=self.engine
            )
            
            if column:
                # Search in specific column
                mask = df[column].astype(str).str.contains(str(search_value), case=False, na=False)
            else:
                # Search in all columns
                mask = df.astype(str).apply(
                    lambda x: x.str.contains(str(search_value), case=False, na=False)
                ).any(axis=1)
            
            results = df[mask].to_dict('records')
            logger.info(f"Found {len(results)} matching rows")
            
            return results
            
        except Exception as e:
            logger.error(f"Error searching: {e}")
            raise
    
    def close(self):
        """Close the workbook and release resources"""
        if self._workbook:
            self._workbook.close()
            self._workbook = None
        
        if self._excel_file:
            self._excel_file.close()
            self._excel_file = None
        
        logger.debug("ExcelReader resources released")


# Example usage and testing
if __name__ == "__main__":
    # Example 1: Basic usage
    try:
        reader = ExcelReader('C:\\OUProjects\\Ingredient-Application.xlsx')
        
        # Get all sheet names
        print("Sheet names:", reader.get_sheet_names())
        
        # Read first sheet
        data = reader.read_sheet(0)
        print(f"Read {len(data)} rows from first sheet")
        
        # Read specific columns
        data = reader.read_sheet('Ingredients', columns=[ 'Brand Name'])
        print(data[:5])  # Print first 5 rows
        # Search for data
        results = reader.search('FC', column='Brand Name')
        print(f"Found {len(results)} results")
        
    except Exception as e:
        print(f"Error: {e}")
    
    # Example 2: Context manager usage
    with ExcelReader('C:\\OUProjects\\Ingredient-Application.xlsx') as reader:
        sheet_info = reader.get_sheet_info('Ingredients')
        print(f"Sheet info: {sheet_info}")
        column_names = sheet_info['column_names']
        rows = sheet_info['total_rows']
        print(f"Column names: {column_names}")  
        for column in column_names:
            print(f"- {column}")   
            data = reader.read_sheet('Ingredients', columns=[column])
            print(f"2 {rows} rows for column '{column}': {data[:1]}")